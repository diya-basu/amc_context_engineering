"""
taxonomy.py
============
Builds taxonomy.json from data/AMFI/ and data/Sub Classification/.
Expected: csv/xlsx files with columns like scheme_name, amc_name, category,
sub_category, benchmark — column names are sniffed loosely (case-insensitive
substring match), so it tolerates messy AMFI exports.
"""
from __future__ import annotations
import json
from pathlib import Path
from typing import Dict, Set

import config

_COLUMN_HINTS = {
    "fund_houses":        ["amc", "fund house", "asset management"],
    "scheme_names":       ["scheme name", "fund name"],
    "categories":         ["category"],
    "sub_categories":     ["sub category", "sub-category", "subcategory"],
    "benchmarks":         ["benchmark"],
}


def _read_tabular(path: Path):
    if path.suffix.lower() == ".csv":
        import csv
        with open(path, newline="", encoding="utf-8", errors="ignore") as f:
            rows = list(csv.reader(f))
        if not rows:
            return [], []
        return rows[0], rows[1:]
    elif path.suffix.lower() in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [[str(c).strip() if c is not None else "" for c in r]
                for r in ws.iter_rows(values_only=True)]
        if not rows:
            return [], []
        return rows[0], rows[1:]
    return [], []


def _match_columns(header: list[str]) -> Dict[str, int]:
    header_lower = [h.lower() for h in header]
    matched = {}
    for key, hints in _COLUMN_HINTS.items():
        for i, h in enumerate(header_lower):
            if any(hint in h for hint in hints):
                matched[key] = i
                break
    return matched


def build_taxonomy(verbose: bool = True) -> Dict[str, list]:
    taxonomy: Dict[str, Set[str]] = {k: set() for k in _COLUMN_HINTS}

    folders = [config.AMFI_DIR, config.SUBCLASS_DIR]
    for folder in folders:
        if not folder.exists():
            continue
        for path in folder.rglob("*"):
            if path.suffix.lower() not in (".csv", ".xlsx", ".xls"):
                continue
            header, rows = _read_tabular(path)
            if not header:
                continue
            col_map = _match_columns(header)
            for key, col_idx in col_map.items():
                for row in rows:
                    if col_idx < len(row) and row[col_idx].strip():
                        taxonomy[key].add(row[col_idx].strip())
            if verbose:
                print(f"  [taxonomy] {path.name}: matched {list(col_map.keys())}", flush=True)

    result = {k: sorted(v) for k, v in taxonomy.items()}
    config.TAXONOMY_PATH.write_text(json.dumps(result, indent=2, ensure_ascii=False))
    if verbose:
        for k, v in result.items():
            print(f"  [taxonomy] {k}: {len(v)} entries", flush=True)
    return result


def load_taxonomy() -> Dict[str, list]:
    if not config.TAXONOMY_PATH.exists():
        return build_taxonomy()
    return json.loads(config.TAXONOMY_PATH.read_text())


if __name__ == "__main__":
    build_taxonomy()